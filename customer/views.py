from django.shortcuts import render, get_object_or_404
from .models import Customer
from .forms import CustomerForm
from django.contrib import messages
import openpyxl


'''
    Adding new customers to the database
    ans displaying a list of existing customers
'''
def customer_list(request):
    customers = Customer.objects.all()
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            check = Customer.objects.filter(first_name__iexact=first_name, last_name__iexact=last_name)
            if check.exists():
                messages.error(request, "Customer exists already")
            else:
                form.save()
                messages.success(request, "Customer added successfully")
    form = CustomerForm()
    context = {
        'form': form,
        'customers': customers
    }
    return render(request, 'index.html', context)


''' Customer's details with a graph (Chart) '''
def customer_detail(request, slug):
    customer = get_object_or_404(Customer, slug=slug)
    custom_file = customer.file
    
    work_book = openpyxl.load_workbook(custom_file)
    sheets = work_book['Sheet1']
    row = sheets.max_row
    column = sheets.max_column
    income = []
    expense = []

    for i in range(1, row + 1):
        for j in range(1, column):
            income.append(sheets.cell(i, j).value)
        for j in range(2, column + 1):
            expense.append(sheets.cell(i, j).value)
            
    context = {
        'customer': customer,
        'expense' : expense[1:],
        'income' : income[1:]
    }
    return render(request, 'customer_detail.html', context)
